import openpyxl
import json

wb = openpyxl.Workbook()

ws = wb.active
ws['A1'] = 'Title'
ws['B1'] = 'Publisher'
ws['C1'] = 'Publish Time'
ws['D1'] = 'Url'
ws['E1'] = 'Content'

i = 2
with open('result3.json', 'r', encoding='utf-8') as f:
	while True:
		line = f.readline()
		if line:
			dict = json.loads(line)
			ws['A'+str(i)] = dict['title']
			ws['B'+str(i)] = dict['author']
			ws['C'+str(i)] = dict['pub_time']
			ws['D'+str(i)] = dict['url']
			ws['E'+str(i)] = dict['content']
			print('第', i, '条…')
			i += 1
		else:
			break

wb.save('news.xlsx')